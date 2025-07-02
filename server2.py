import os
import datetime
import ping3
import openpyxl
import schedule
import time
import subprocess
import platform
import requests
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Configuration: Your servers list with corrected IP for Server5
servers = [
     {"KETL": "KETL", "ip": "172.24.24.146"},
    {"SMTL": "SMTL", "ip": "172.24.179.113"},
    {"NKTL": "NKTL", "ip": "172.24.134.79"},
    {"LRTL": "LRTL", "ip": "172.24.61.249"},
    {"JVTL": "JVTL", "ip": "172.24.48.115"},
    {"MKTL-15": "MKTL-15", "ip": "172.24.15.161"},
    {"MKTL-16": "MKTL-16", "ip": "172.24.137.166"},
    {"JLTL": "JLTL", "ip": "172.24.116.111"},
]

# Directory to save Excel files
LOG_DIR = "server_logs"
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

# Slack Webhook URL (replace with your own)
SLACK_WEBHOOK_URL = "your_slack_webhook_url"

# Track last status, down time, up time, and notification sent for each server
server_status_tracker = {
    server[next(iter(server))]: {
        "last_status": None,
        "last_down_time": None,
        "last_up_time": None,
        "notification_sent": False  # Track if Down notification was sent
    } for server in servers
}

def get_excel_filename():
    """Generate Excel filename based on current date."""
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    return os.path.join(LOG_DIR, f"server_status_{current_date}.xlsx")

def send_slack_message(server_name, ip, current_time):
    """Send Slack message when a server is Down."""
    payload = {
        "text": f"🚨 Server Down Alert 🚨\nServer: {server_name}\nIP: {ip}\nStatus: Down\nTime: {current_time}"
    }
    try:
        response = requests.post(SLACK_WEBHOOK_URL, json=payload)
        if response.status_code == 200:
            print(f"Slack message sent for {server_name} ({ip})")
        else:
            print(f"Slack message error for {server_name} ({ip}): {response.status_code} - {response.text}")
    except Exception as e:
        print(f"Slack message error for {server_name} ({ip}): {e}")

def ping_server(ip):
    """Ping a server IP with double-check mechanism and return status (Up/Down)."""
    def try_ping():
        # Try ping3 first
        try:
            response_time = ping3.ping(ip, timeout=5)  # 5-second timeout
            if response_time is not None:
                print(f"Ping3 success for {ip}: Response time {response_time}ms")
                return True
            else:
                print(f"Ping3 failed for {ip}: No response")
        except Exception as e:
            print(f"Ping3 error for {ip}: {e}")
        
        # Fallback to system ping
        try:
            param = '-n' if platform.system().lower() == 'windows' else '-c'
            result = subprocess.run(
                ['ping', param, '1', ip],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=10
            )
            if result.returncode == 0:
                print(f"System ping success for {ip}: {result.stdout}")
                return True
            else:
                print(f"System ping failed for {ip}: {result.stderr}")
                return False
        except Exception as e:
            print(f"System ping error for {ip}: {e}")
            return False
    
    # Initial ping attempt
    if try_ping():
        return "Up"
    
    # Wait 120 seconds and retry if initial ping failed
    print(f"Initial ping failed for {ip}, waiting 120 seconds for second check...")
    time.sleep(120)
    
    # Second ping attemptprint("DEBUG: Raw MONITOR_IDS =", os.getenv("MONITOR_IDS"))
    if try_ping():
        return "Up"
    return "Down"

def initialize_excel_sheet(workbook, sheet_name):
    """Initialize a new sheet with updated headers if it doesn't exist."""
    # Sanitize sheet name to avoid invalid characters
    sheet_name = sheet_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        # Updated headers
        headers = ["Date & Time", "Server IP", "Server Status", "Last Down Time", "Last Up Time", "Downtime Duration (Min)"]
        for col_num, header in enumerate(headers, 1):
            sheet[f"{get_column_letter(col_num)}1"] = header
    return workbook[sheet_name]

def log_server_status():
    """Check server status and log all fields to Excel every 5 minutes."""
    print("Servers configuration:", servers)
    excel_file = get_excel_filename()
    
    # Load or create Excel workbook
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except Exception as e:
        print(f"Error loading Excel file {excel_file}: {e}. Creating new workbook.")
        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data_logged = False

    # Define colors for Server Status column
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for server in servers:
        try:
            server_key = next(iter(server))
            if server_key == "ip":
                server_key = next(iter(server.keys() - {"ip"}))
            server_name = server[server_key]
            ip = server["ip"]
        except (KeyError, StopIteration) as e:
            print(f"Error: Invalid server entry {server}: {e}")
            continue
        
        status = ping_server(ip)
        down_time = ""
        up_time = ""
        duration = ""
        
        # Get current timestamp as datetime object for calculations
        current_time_dt = datetime.datetime.strptime(current_time, "%Y-%m-%d %H:%M:%S")
        
        # Track status changes
        tracker = server_status_tracker[server_key]
        last_status = tracker["last_status"]
        
        if status == "Down" and not tracker["notification_sent"]:
            message = f"🚨 Server Down Alert 🚨\nServer: {server_name}\nIP: {ip}\nTime: {current_time}"
            send_whatsapp_alert("918328618110", message)  # Replace with full international number
            send_slack_message(server_name, ip, current_time)
            tracker["notification_sent"] = True
        # if status == "Down":
        #     # If server is Down and wasn't Down before, set Down Time and send notification
        #     if last_status != "Down":
        #         tracker["last_down_time"] = current_time_dt
        #         tracker["notification_sent"] = False  # Reset notification flag
        #     # Use the persisted Down Time
        #     if tracker["last_down_time"]:
        #         down_time = tracker["last_down_time"].strftime("%Y-%m-%d %H:%M:%S")
        #     # Send Slack message if not already sent
        #     if not tracker["notification_sent"]:
        #         send_slack_message(server_name, ip, current_time)
        #         tracker["notification_sent"] = True
        elif status == "Up":
            # If server came back up
            if last_status == "Down":
                tracker["last_up_time"] = current_time_dt
                up_time = current_time
                if tracker["last_down_time"]:
                    # Calculate duration
                    duration_dt = current_time_dt - tracker["last_down_time"]
                    duration = round(duration_dt.total_seconds() / 60, 2)  # Duration in minutes
                    tracker["last_down_time"] = None  # Reset down time
                tracker["notification_sent"] = False  # Reset notification flag
            # Always log Up Time if available
            if tracker["last_up_time"]:
                up_time = tracker["last_up_time"].strftime("%Y-%m-%d %H:%M:%S")
        
        tracker["last_status"] = status
        
        # Get or create sheet for the server
        sheet = initialize_excel_sheet(workbook, server_name)
        
        # Append all fields to Excel
        row = [current_time, ip, status, down_time, up_time, duration]
        sheet.append(row)
        data_logged = True
        
        # Apply color to Server Status cell
        status_cell = f"C{sheet.max_row}"  # Server Status column is C
        sheet[status_cell].fill = red_fill if status == "Down" else green_fill
        
        # Print to terminal based on status
        if status == "Down":
            print(f"Logged: {server_name} ({ip}) - Down at {current_time}" + 
                  (f", Last Down Time: {down_time}" if down_time else ""))
        else:
            print(f"Logged: {server_name} ({ip}) - Up at {current_time}" + 
                  (f", Last Up Time: {up_time}" if up_time else "") + 
                  (f", Downtime Duration: {duration} min" if duration else ""))

    # Save the workbook only if data was logged
    if data_logged:
        try:
            workbook.save(excel_file)
            print(f"Saved status to {excel_file}\n")
        except Exception as e:
            print(f"Error saving Excel file: {e}")
    else:
        print("No data logged, skipping Excel save.")

#whatsapp notifier function
def send_whatsapp_alert(number, message):
    """Send a WhatsApp message using local Node.js notifier."""
    try:
        res = requests.post("http://localhost:3000/send", json={
            "number": number,
            "message": message
        })
        if res.status_code == 200:
            print("WhatsApp alert sent.")
        else:
            print(f"WhatsApp alert failed: {res.text}")
    except Exception as e:
        print(f"Error sending WhatsApp alert: {e}")


def main():
    """Main function to schedule and run the status check every 5 minutes."""
    schedule.every(5).minutes.do(log_server_status)
    
    print("Starting server status monitoring...")
    log_server_status()
    
    while True:
        try:
            schedule.run_pending()
            time.sleep(1)
        except KeyboardInterrupt:
            print("\nStopping server status monitoring...")
            break
        except Exception as e:
            print(f"Unexpected error: {e}")
            time.sleep(60)

if __name__ == "__main__":
    main()
    